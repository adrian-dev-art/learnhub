from django.shortcuts import render, redirect, get_object_or_404
from django.db import models
from django.conf import settings
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from .models import Course, Module, Enrollment, Assessment, AssessmentResult, Certificate, User, Commission, CommissionRate
from .forms import CustomUserCreationForm, ProfileForm, AssessmentSubmissionForm, CourseForm, ModuleForm, AssessmentForm
from .utils import send_access_key_email, execute_python_code, generate_certificate_pdf, parse_docx_to_modules
import tempfile
import os
from django.core.exceptions import PermissionDenied
from functools import wraps
import json
import markdown
from decimal import Decimal
from django.core.files.storage import FileSystemStorage
from datetime import datetime


def mentor_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if request.user.is_authenticated and (request.user.role == 'penulis' or request.user.is_mentor):
            return view_func(request, *args, **kwargs)
        raise PermissionDenied
    return _wrapped_view


def admin_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if request.user.is_authenticated and request.user.role == 'admin':
            return view_func(request, *args, **kwargs)
        raise PermissionDenied
    return _wrapped_view


def owner_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if request.user.is_authenticated and request.user.role == 'owner':
            return view_func(request, *args, **kwargs)
        raise PermissionDenied
    return _wrapped_view


def home(request):
    """Home page"""
    try:
        # Get featured courses with modules (limit to 6)
        courses = Course.objects.all()[:6]
        # Attach modules to each course
        for course in courses:
            course.module_list = course.modules.all()[:5]  # First 5 modules
            course.total_modules = course.modules.count()
    except Exception as e:
        print(f"Database error: {e}")
        courses = []
    return render(request, 'landing/home.html', {'courses': courses})


def catalog(request):
    """Course catalog view"""
    try:
        courses = Course.objects.all()
    except Exception as e:
        print(f"Database error: {e}")
        courses = []
    return render(request, 'student/catalog.html', {'courses': courses})


@login_required
def student_dashboard(request):
    """Student dashboard with owned courses and recommendations"""
    try:
        # Get owned courses (completed payments)
        enrollments = Enrollment.objects.filter(
            user=request.user,
            payment_status='completed'
        )

        # Get all courses for catalog/recommendations
        all_courses = Course.objects.all()

        # Dummy data for "Best Seller" and "Related" for now
        best_sellers = all_courses.order_by('-created_at')[:3]
        related_courses = all_courses.order_by('-created_at')[3:6]  # Different set of courses

        context = {
            'enrollments': enrollments,
            'best_sellers': best_sellers,
            'related_courses': related_courses,
            'all_courses': all_courses
        }
    except Exception as e:
        print(f"Dashboard error: {e}")
        context = {
            'enrollments': [],
            'best_sellers': [],
            'related_courses': [],
            'all_courses': []
        }

    return render(request, 'student/dashboard.html', context)


def course_detail(request, course_id):
    """Course detail and enrollment page"""
    course = get_object_or_404(Course, id=course_id)
    modules = course.modules.all()
    first_module = modules.first()

    # Check if user already enrolled (only if logged in)
    enrollment = None
    progress_percentage = 0
    certificate = None

    if request.user.is_authenticated:
        enrollment = Enrollment.objects.filter(
            user_id=request.user.id,
            course_id=course.id,
            payment_status='completed'
        ).first()

        if enrollment:
            # Calculate progress
            completed_count = len(enrollment.progress.get('completed_modules', []))
            total_count = modules.count()
            if total_count > 0:
                progress_percentage = int((completed_count / total_count) * 100)

            # Check for certificate
            try:
                certificate = Certificate.objects.get(enrollment_id=enrollment.id)
            except Certificate.DoesNotExist:
                certificate = None

    context = {
        'course': course,
        'modules': modules,
        'first_module': first_module,
        'enrollment': enrollment,
        'progress_percentage': progress_percentage,
        'certificate': certificate,
        'already_enrolled': enrollment is not None
    }
    return render(request, 'student/course_detail.html', context)


def payment(request, course_id):
    """Payment and profile completion page - handles both new and existing users"""
    course = get_object_or_404(Course, id=course_id)

    # Check if already enrolled (for logged-in users)
    if request.user.is_authenticated:
        existing_enrollment = Enrollment.objects.filter(
            user_id=request.user.id,
            course_id=course.id,
            payment_status='completed'
        ).first()

        if existing_enrollment:
            messages.info(request, 'You are already enrolled in this course.')
            return redirect('course_viewer', enrollment_id=existing_enrollment.id)

    if request.method == 'POST':
        # Handle new user registration + payment
        if not request.user.is_authenticated:
            # Get form data
            username = request.POST.get('username')
            email = request.POST.get('email')
            password1 = request.POST.get('password1')
            password2 = request.POST.get('password2')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            phone = request.POST.get('phone', '')

            # Validate
            if not all([username, email, password1, password2, first_name, last_name]):
                messages.error(request, 'Please fill in all required fields.')
                return render(request, 'student/payment.html', {'course': course})

            if password1 != password2:
                messages.error(request, 'Passwords do not match.')
                return render(request, 'student/payment.html', {'course': course})

            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists.')
                return render(request, 'student/payment.html', {'course': course})

            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email already registered.')
                return render(request, 'student/payment.html', {'course': course})

            # Create user
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password1,
                first_name=first_name,
                last_name=last_name,
                phone=phone,
                profile_completed=True
            )

            # Log the user in
            login(request, user)

        else:
            # Existing user - skip profile update as requested
            user = request.user
            # Optional: ensure profile_completed is True if it wasn't already
            if not user.profile_completed:
                user.profile_completed = True
                user.save()

        # Create enrollment for both new and existing users
        enrollment = Enrollment.objects.create(
            user_id=request.user.id,
            course_id=course.id,
            payment_status='completed'  # Demo mode
        )

        # Send access key via email
        send_access_key_email(
            request.user.email,
            course.title,
            enrollment.access_key
        )

        messages.success(
            request,
            f'Enrollment successful! Your access key has been sent to {request.user.email}'
        )
        return redirect('course_viewer', enrollment_id=enrollment.id)

    # GET request
    profile_form = None
    if request.user.is_authenticated:
        profile_form = ProfileForm(instance=request.user)

    context = {
        'course': course,
        'profile_form': profile_form
    }
    return render(request, 'student/payment.html', context)


@login_required
def course_viewer(request, enrollment_id):
    """Interactive course viewer with sequential navigation and Markdown support"""
    is_preview = request.GET.get('preview') == 'true'

    if enrollment_id == 0 and is_preview:
        # Free preview mode for guests or non-enrolled users
        course_id = request.GET.get('course_id')
        course = get_object_or_404(Course, id=course_id)
        enrollment = None
        modules = course.modules.all()
        current_module = modules.first()
        completed = []
    else:
        # Normal enrollment mode
        enrollment = get_object_or_404(Enrollment, id=enrollment_id, user_id=request.user.id)
        course = enrollment.course
        modules = course.modules.all()

        if not enrollment.progress:
            enrollment.progress = {'completed_modules': []}
            enrollment.save()

        completed = enrollment.progress.get('completed_modules', [])

        # Find the first incomplete module (the "current" one in sequence)
        first_incomplete = None
        for m in modules:
            if m.id not in completed:
                first_incomplete = m
                break

        # Handle specific module request
        requested_module_id = request.GET.get('module_id')
        if requested_module_id:
            try:
                requested_module_id = int(requested_module_id)
                # Navigation Logic: Allow if completed OR if it's the next incomplete module
                if requested_module_id in completed or (first_incomplete and requested_module_id == first_incomplete.id):
                    current_module = get_object_or_404(Module, id=requested_module_id, course_id=course.id)
                else:
                    messages.warning(request, "This module is locked. Please complete the previous modules first.")
                    current_module = first_incomplete or modules.last()
            except (ValueError, TypeError):
                current_module = first_incomplete or modules.last()
        else:
            # Default to first incomplete
            current_module = first_incomplete or modules.last()

    if not current_module and not is_preview:
        messages.error(request, "This course has no modules yet.")
        return redirect('student_dashboard')

    # Annotate modules with status for the template
    for m in modules:
        m.is_completed = m.id in completed
        m.is_current = current_module and m.id == current_module.id

    # Render Markdown content for the current module
    if current_module and current_module.content:
        # Convert markdown to HTML
        current_module.content_html = markdown.markdown(
            current_module.content,
            extensions=['extra', 'codehilite', 'toc', 'fenced_code']
        )
    else:
        if current_module:
            current_module.content_html = ""

    context = {
        'enrollment': enrollment,
        'course': course,
        'modules': modules,
        'current_module': current_module,
        'completed_modules': completed,
        'first_incomplete': first_incomplete,
        'is_preview': is_preview,
        'saved_quiz_data': enrollment.progress.get('quiz_data', {}).get(str(current_module.id), {}) if enrollment and current_module else {}
    }
    return render(request, 'student/course_viewer.html', context)


@login_required
@require_POST
def mark_module_complete(request, enrollment_id, module_id):
    """Mark a module as completed"""
    enrollment = get_object_or_404(Enrollment, id=enrollment_id, user_id=request.user.id)
    module = get_object_or_404(Module, id=module_id, course=enrollment.course)

    if not enrollment.progress:
        enrollment.progress = {'completed_modules': []}

    completed = enrollment.progress.get('completed_modules', [])
    if module_id not in completed:
        completed.append(module_id)
        enrollment.progress['completed_modules'] = completed
        enrollment.save()

    return JsonResponse({'success': True, 'completed': completed})


@login_required
@require_POST
def save_module_quiz_progress(request, enrollment_id, module_id):
    """Save user's progress on a module quiz"""
    enrollment = get_object_or_404(Enrollment, id=enrollment_id, user_id=request.user.id)

    try:
        data = json.loads(request.body)
        answers = data.get('answers')
        score = data.get('score')

        if not enrollment.progress:
            enrollment.progress = {'completed_modules': []}

        # Initialize quiz_data if not exists
        if 'quiz_data' not in enrollment.progress:
            enrollment.progress['quiz_data'] = {}

        # Save quiz result for this module
        enrollment.progress['quiz_data'][str(module_id)] = {
            'answers': answers,
            'score': score,
            'completed_at': str(datetime.now())
        }

        enrollment.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
@require_POST
def execute_code(request):
    """Execute Python code from the interactive compiler"""
    try:
        data = json.loads(request.body)
        code = data.get('code', '')
        inputs = data.get('inputs', '')

        if not code:
            return JsonResponse({'success': False, 'error': 'No code provided'})

        success, output, error = execute_python_code(code, inputs)

        return JsonResponse({
            'success': success,
            'output': output,
            'error': error
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def assessment_view(request, enrollment_id):
    """Course assessment view"""
    enrollment = get_object_or_404(Enrollment, id=enrollment_id, user_id=request.user.id)
    course = enrollment.course

    try:
        assessment = course.assessment
    except Assessment.DoesNotExist:
        messages.error(request, 'No assessment available for this course.')
        return redirect('course_viewer', enrollment_id=enrollment_id)

    # Check if already completed
    existing_result = AssessmentResult.objects.filter(
        user_id=request.user.id,
        assessment_id=assessment.id,
        passed=True
    ).first()

    if existing_result:
        messages.info(request, 'You have already passed this assessment.')
        return redirect('certificate_view', enrollment_id=enrollment_id)

    # Get questions (prefer JSON if exists)
    is_json = False
    if assessment.questions_json:
        questions = assessment.questions_json
        is_json = True
    else:
        questions = assessment.questions.all().prefetch_related('choices')

    if request.method == 'POST':
        form = AssessmentSubmissionForm(request.POST, questions=questions)

        if form.is_valid():
            # Calculate score
            correct = 0
            answers = {}
            
            for i, question in enumerate(questions):
                if is_json:
                    q_id = i
                    user_choice_idx = form.cleaned_data.get(f'question_{q_id}')
                    answers[f'question_{q_id}'] = user_choice_idx
                    if user_choice_idx is not None:
                        try:
                            idx = int(user_choice_idx)
                            if 0 <= idx < len(question['options']):
                                opt = question['options'][idx]
                                user_answer_text = opt['text'] if isinstance(opt, dict) else opt
                                if user_answer_text == question.get('correct_answer'):
                                    correct += 1
                        except (ValueError, TypeError):
                            pass
                else:
                    q_id = question.id
                    user_choice_id = form.cleaned_data.get(f'question_{q_id}')
                    answers[f'question_{q_id}'] = user_choice_id
                    if user_choice_id:
                        try:
                            choice = question.choices.get(id=user_choice_id)
                            if choice.is_correct:
                                correct += 1
                        except Exception:
                            pass

            total_questions = len(questions) if is_json else questions.count()
            if total_questions > 0:
                score = int((correct / total_questions) * 100)
            else:
                score = 0

            passed = score >= assessment.passing_score

            # Save result
            result = AssessmentResult.objects.create(
                user_id=request.user.id,
                assessment_id=assessment.id,
                enrollment_id=enrollment.id,
                score=score,
                answers=answers,
                passed=passed
            )

            if passed:
                # Mark enrollment as completed
                enrollment.completed = True
                enrollment.save()

                # Generate certificate
                Certificate.objects.get_or_create(
                    user_id=request.user.id,
                    course_id=course.id,
                    enrollment_id=enrollment.id
                )

                messages.success(request, f'Congratulations! You passed with {score}%')
                return redirect('certificate_view', enrollment_id=enrollment_id)
            else:
                messages.error(
                    request,
                    f'You scored {score}%. You need {assessment.passing_score}% to pass. Please try again.'
                )
    else:
        form = AssessmentSubmissionForm(questions=questions)

    # Attach form fields and IDs for template rendering
    rendered_questions = []
    for i, question in enumerate(questions):
        if is_json:
            q_id = i
            q_text = question.get('question', '')
            q_image_url = question.get('image_url', '')
            # For JSON, options might be strings or objects
            q_options = []
            for idx, opt in enumerate(question.get('options', [])):
                if isinstance(opt, dict):
                    q_options.append({
                        'id': idx, 
                        'text': opt.get('text', ''), 
                        'image_url': opt.get('image_url', '')
                    })
                else:
                    q_options.append({'id': idx, 'text': opt, 'image_url': ''})
        else:
            q_id = question.id
            q_text = question.text
            q_image_url = question.image.url if hasattr(question, 'image') and question.image else ''
            # Map choice models to standard dict
            q_options = []
            for choice in question.choices.all():
                q_options.append({
                    'id': choice.id,
                    'text': choice.text,
                    'image_url': choice.image.url if hasattr(choice, 'image') and choice.image else ''
                })

        field_name = f'question_{q_id}'
        form_field = form[field_name] if field_name in form.fields else None
        
        rendered_questions.append({
            'id': q_id,
            'text': q_text,
            'image_url': q_image_url,
            'options': q_options,
            'form_field': form_field
        })

    context = {
        'enrollment': enrollment,
        'assessment': assessment,
        'questions': rendered_questions,
        'form': form
    }
    return render(request, 'student/assessment.html', context)


@login_required
def certificate_view(request, enrollment_id):
    """View and download certificate"""
    enrollment = get_object_or_404(Enrollment, id=enrollment_id, user_id=request.user.id)

    try:
        certificate = Certificate.objects.get(enrollment_id=enrollment.id)
    except Certificate.DoesNotExist:
        messages.error(request, 'Certificate not available yet.')
        return redirect('course_viewer', enrollment_id=enrollment_id)

    context = {
        'certificate': certificate,
        'enrollment': enrollment
    }
    return render(request, 'student/certificate.html', context)


@login_required
def download_certificate(request, certificate_id):
    """Download certificate as PDF"""
    certificate = get_object_or_404(Certificate, id=certificate_id, user_id=request.user.id)

    # Generate PDF
    pdf_buffer = generate_certificate_pdf(certificate)

    # Create response
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="certificate_{certificate.certificate_id}.pdf"'

    return response


# Authentication views
def register(request):
    """User registration"""
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Registration successful!')
            return redirect('catalog')
    else:
        form = CustomUserCreationForm()

    return render(request, 'landing/register.html', {'form': form})


def login_view(request):
    """User login"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            messages.success(request, 'Login successful!')
            next_url = request.GET.get('next', 'catalog')
            return redirect(next_url)
        else:
            messages.error(request, 'Invalid username or password.')

    return render(request, 'landing/login.html')


def logout_view(request):
    """User logout"""
    logout(request)
    messages.success(request, 'Logged out successfully!')
    return redirect('home')


@login_required
@mentor_required
def mentor_dashboard(request):
    """Mentor dashboard overview"""
    # Get courses mentored by the user
    mentor_courses = Course.objects.filter(mentor=request.user)

    # Stats
    total_courses = mentor_courses.count()

    # Get enrollments for mentor's courses
    mentor_enrollments = Enrollment.objects.filter(course__mentor=request.user)
    total_enrollments = mentor_enrollments.count()
    completed_enrollments = mentor_enrollments.filter(completed=True).count()

    # Get total commission from Commission model
    commissions = Commission.objects.filter(user=request.user)
    total_commission = commissions.aggregate(total=models.Sum('amount'))['total'] or Decimal('0')

    context = {
        'courses': mentor_courses,
        'total_courses': total_courses,
        'total_enrollments': total_enrollments,
        'completed_enrollments': completed_enrollments,
        'total_commission': total_commission,
    }
    return render(request, 'mentor/dashboard.html', context)


@login_required
@mentor_required
def mentor_commission_detail(request):
    """Detailed view of commission earnings"""
    commissions = Commission.objects.filter(user=request.user).select_related('course', 'enrollment__user').order_by('-created_at')

    total_commission = commissions.aggregate(total=models.Sum('amount'))['total'] or Decimal('0')

    context = {
        'commissions': commissions,
        'total_commission': total_commission
    }
    return render(request, 'mentor/commission_detail.html', context)


@login_required
@admin_required
def admin_dashboard(request):
    """Admin dashboard - redirects to Django admin"""
    return redirect('/admin/')


@login_required
@owner_required
def owner_dashboard(request):
    """Owner dashboard overview"""
    # Calculate financial stats from Commission model
    # Total Revenue = Sum of all completed enrollment course prices
    all_enrollments = Enrollment.objects.filter(payment_status='completed')
    total_revenue = Decimal('0')
    for en in all_enrollments:
        price = en.course.price
        if hasattr(price, 'to_decimal'):
            price = price.to_decimal()
        elif not isinstance(price, Decimal):
            price = Decimal(str(price))
        total_revenue += price

    # Platform Profit = Sum of 'admin' and 'layanan' commissions
    platform_profit = Commission.objects.filter(
        role__in=['admin', 'layanan']
    ).aggregate(total=models.Sum('amount'))['total'] or Decimal('0')

    # Top courses
    courses = Course.objects.all()
    course_list = []
    for c in courses:
        enrollment_count = Enrollment.objects.filter(course=c, payment_status='completed').count()
        course_list.append({
            'title': c.title,
            'enrollment_count': enrollment_count,
            'mentor': c.mentor.get_full_name() or c.mentor.username
        })

    top_courses = sorted(course_list, key=lambda x: x['enrollment_count'], reverse=True)[:5]
    recent_enrollments = all_enrollments.order_by('-enrolled_at')[:10]

    context = {
        'total_revenue': total_revenue,
        'platform_profit': platform_profit,
        'top_courses': top_courses,
        'recent_enrollments': recent_enrollments,
    }
    return render(request, 'owner/dashboard.html', context)



@login_required
def schedule(request):
    """Schedule view - Placeholder"""
    return render(request, 'student/schedule.html')


@login_required
def settings_view(request):
    """User settings and profile update"""
    if request.method == 'POST':
        form = ProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully.')
            return redirect('settings')
    else:
        form = ProfileForm(instance=request.user)

    return render(request, 'student/settings.html', {'form': form})


@login_required
def progress_view(request):
    """Detailed progress and achievements view"""
    enrollments = Enrollment.objects.filter(
        user=request.user,
        payment_status='completed'
    )

    # Calculate stats
    total_enrolled = enrollments.count()
    completed_courses = 0
    certificates = Certificate.objects.filter(user=request.user)

    active_courses = []
    completed_list = []

    for en in enrollments:
        if en.completed:
            completed_courses += 1
            completed_list.append(en)
        else:
            active_courses.append(en)

    context = {
        'enrollments': enrollments,
        'total_enrolled': total_enrolled,
        'completed_courses': completed_courses,
        'active_courses': active_courses,
        'completed_list': completed_list,
        'certificates': certificates
    }
    return render(request, 'student/progress.html', context)



@login_required
@mentor_required
def mentor_course_add(request):
    """Create a new course"""
    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES)
        if form.is_valid():
            course = form.save(commit=False)
            course.mentor = request.user
            course.save()
            messages.success(request, f'Course "{course.title}" created successfully.')
            return redirect('mentor_dashboard')
    else:
        form = CourseForm()

    return render(request, 'mentor/course_form.html', {'form': form, 'action': 'Create'})


@login_required
@mentor_required
def mentor_course_edit(request, course_id):
    """Edit an existing course"""
    course = get_object_or_404(Course, id=course_id, mentor=request.user)

    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, f'Course "{course.title}" updated successfully.')
            return redirect('mentor_dashboard')
    else:
        form = CourseForm(instance=course)

    return render(request, 'mentor/course_form.html', {'form': form, 'course': course, 'action': 'Edit'})


@login_required
@mentor_required
def mentor_course_detail(request, course_id):
    """Detailed view of a course for the mentor"""
    course = get_object_or_404(Course, id=course_id, mentor=request.user)
    modules = course.modules.all()
    enrollments = Enrollment.objects.filter(course=course)

    # Calculate progress for each enrollment
    module_count = modules.count()
    for enrollment in enrollments:
        completed_count = len(enrollment.progress.get('completed_modules', []))
        if module_count > 0:
            enrollment.calculated_progress = int((completed_count / module_count) * 100)
        else:
            enrollment.calculated_progress = 0

    context = {
        'course': course,
        'modules': modules,
        'enrollments': enrollments,
    }
    return render(request, 'mentor/course_detail.html', context)


@login_required
@mentor_required
def mentor_course_import_doc(request, course_id):
    """Import course modules from a .docx file"""
    course = get_object_or_404(Course, id=course_id, mentor=request.user)

    if request.method == 'POST' and request.FILES.get('doc_file'):
        doc_file = request.FILES['doc_file']

        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as tmp:
            for chunk in doc_file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            modules_data = parse_docx_to_modules(tmp_path)

            # Get current max order
            last_module = course.modules.order_by('order').last()
            current_order = (last_module.order + 10) if last_module else 10

            created_count = 0
            for m_data in modules_data:
                Module.objects.create(
                    course=course,
                    title=m_data['title'],
                    content=m_data['content'],
                    content_type=m_data['content_type'],
                    order=current_order
                )
                current_order += 10
                created_count += 1

            messages.success(request, f'Successfully imported {created_count} modules from document.')
            return redirect('mentor_course_detail', course_id=course.id)

        except Exception as e:
            messages.error(request, f'Error importing document: {str(e)}')
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    return render(request, 'mentor/import_doc.html', {'course': course})


@login_required
@mentor_required
def import_module_content(request):
    """Import content from a docx file for a single module"""
    if request.method == 'POST' and request.FILES.get('doc_file'):
        doc_file = request.FILES['doc_file']

        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as tmp:
            for chunk in doc_file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            from .utils import convert_docx_to_html
            html_content = convert_docx_to_html(tmp_path)
            return JsonResponse({'success': True, 'content': html_content})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    return JsonResponse({'success': False, 'error': 'Invalid request'})


@login_required
@mentor_required
def mentor_module_add(request, course_id):
    """Add a new module to a course"""
    course = get_object_or_404(Course, id=course_id, mentor=request.user)

    if request.method == 'POST':
        form = ModuleForm(request.POST, request.FILES)
        if form.is_valid():
            module = form.save(commit=False)
            module.course = course
            module.save()
            messages.success(request, f'Module "{module.title}" added successfully.')
            return redirect('mentor_course_detail', course_id=course.id)
    else:
        # Get the next order number
        last_module = course.modules.order_by('order').last()
        next_order = (last_module.order + 10) if last_module else 10
        form = ModuleForm(initial={'order': next_order})

    return render(request, 'mentor/module_form.html', {'form': form, 'course': course, 'action': 'Add'})


@login_required
@mentor_required
def mentor_module_edit(request, module_id):
    """Edit an existing module"""
    module = get_object_or_404(Module, id=module_id, course__mentor=request.user)
    course = module.course

    if request.method == 'POST':
        form = ModuleForm(request.POST, request.FILES, instance=module)
        if form.is_valid():
            form.save()
            messages.success(request, f'Module "{module.title}" updated successfully.')
            return redirect('mentor_course_detail', course_id=course.id)
    else:
        form = ModuleForm(instance=module)

    return render(request, 'mentor/module_form.html', {'form': form, 'module': module, 'course': course, 'action': 'Edit'})


@login_required
@mentor_required
def mentor_assessment_edit(request, course_id):
    """Create or edit course assessment"""
    course = get_object_or_404(Course, id=course_id, mentor=request.user)

    try:
        assessment = course.assessment
    except Assessment.DoesNotExist:
        assessment = None

    if request.method == 'POST':
        form = AssessmentForm(request.POST, instance=assessment)
        if form.is_valid():
            assessment = form.save(commit=False)
            assessment.course = course

            # Handle questions from POST data
            questions = []
            
            # Find all question indices present in POST to handle gaps
            q_indices = []
            for key in request.POST.keys():
                if key.startswith('question_text_'):
                    try:
                        q_indices.append(int(key.split('_')[-1]))
                    except (ValueError, IndexError):
                        pass
            
            q_indices.sort()
            
            for q_idx in q_indices:
                q_text = request.POST.get(f'question_text_{q_idx}')

                # Get options
                options = []
                o_idx = 0
                while f'option_text_{q_idx}_{o_idx}' in request.POST:
                    opt_text = request.POST.get(f'option_text_{q_idx}_{o_idx}')
                    
                    # Handle option image
                    opt_image_url = request.POST.get(f'existing_option_image_{q_idx}_{o_idx}', '')
                    opt_file = request.FILES.get(f'option_image_{q_idx}_{o_idx}')
                    
                    if opt_file:
                        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'assessments'))
                        filename = fs.save(opt_file.name, opt_file)
                        opt_image_url = os.path.join(settings.MEDIA_URL, 'assessments', filename).replace('\\', '/')
                    
                    if opt_text: 
                        options.append({
                            'text': opt_text,
                            'image_url': opt_image_url
                        })
                    o_idx += 1

                # Get correct answer index
                correct_idx_str = request.POST.get(f'correct_index_{q_idx}')
                if correct_idx_str and options:
                    try:
                        correct_idx = int(correct_idx_str)
                        if 0 <= correct_idx < len(options):
                            correct_answer = options[correct_idx]['text']
                        else:
                            correct_answer = options[0]['text']
                    except ValueError:
                        correct_answer = options[0]['text']
                elif options:
                    correct_answer = options[0]['text']
                else:
                    correct_answer = ""

                # Get or keep question image
                image_url = request.POST.get(f'existing_image_{q_idx}', '')
                q_image = request.FILES.get(f'question_image_{q_idx}')
                
                if q_image:
                    fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'assessments'))
                    filename = fs.save(q_image.name, q_image)
                    image_url = os.path.join(settings.MEDIA_URL, 'assessments', filename).replace('\\', '/')

                if q_text and options:
                    questions.append({
                        'question': q_text,
                        'options': options,
                        'correct_answer': correct_answer,
                        'image_url': image_url
                    })
                q_idx += 1

            assessment.questions_json = questions
            assessment.save()
            messages.success(request, 'Assessment saved successfully.')
            return redirect('mentor_course_detail', course_id=course.id)
    else:
        form = AssessmentForm(instance=assessment)
    return render(request, 'mentor/assessment_form.html', {
        'form': form,
        'course': course,
        'assessment': assessment,
        'questions_json': json.dumps(assessment.questions_json) if assessment and assessment.questions_json else '[]'
    })


@login_required
@mentor_required
def download_assessment_template(request):
    """Download Excel template for assessment questions"""
    import openpyxl
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assessment Questions"

    # Headers
    headers = ['Question', 'Option 1', 'Option 2', 'Option 3', 'Option 4', 'Correct Option (1-4)']
    ws.append(headers)

    # Example row
    example = ['What is the capital of France?', 'London', 'Berlin', 'Paris', 'Madrid', '3']
    ws.append(example)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=assessment_template.xlsx'
    wb.save(response)
    return response


@login_required
@mentor_required
@require_POST
def import_assessment_questions(request):
    """Import assessment questions from Excel"""
    import openpyxl

    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No file uploaded'})

    file = request.FILES['file']

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        questions = []

        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: # Skip empty rows
                continue

            question_text = str(row[0]).strip()
            options = []

            # Get options (columns 1-4)
            for i in range(1, 5):
                if i < len(row) and row[i]:
                    options.append(str(row[i]).strip())

            # Get correct answer index (column 5) - 1-based index in Excel to 0-based
            correct_idx = 0
            if len(row) > 5 and row[5]:
                try:
                    val = int(row[5])
                    if 1 <= val <= len(options):
                        correct_idx = val - 1
                except ValueError:
                    pass

            if question_text and len(options) >= 2:
                questions.append({
                    'question': question_text,
                    'options': options,
                    'correct_answer': options[correct_idx] if options else '',
                    'correct_index': correct_idx # Helper for frontend
                })

        return JsonResponse({'success': True, 'questions': questions})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
